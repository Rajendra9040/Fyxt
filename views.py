import os, re, glob, datetime, pytz
import pandas as pd
from docx2pdf import convert
from pdf2docx import Converter
from PyPDF2 import PdfFileWriter, PdfFileReader
from openpyxl import load_workbook

from django.utils import timezone

from django.db import transaction
from django.db.models import Value, Q
from django.db.models.functions import Concat

from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache

from django.conf import settings
from django.core.files import File
from django.core.files.storage import get_storage_class
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.contrib.postgres.search import SearchVector

from django.shortcuts import render
from django.shortcuts import get_object_or_404

from django_filters.rest_framework import DjangoFilterBackend

from rest_framework import status, permissions, viewsets, filters
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.decorators import action

from codequick.utils.mixins import SerializerClassMixin

from cqdashboard.models import *
from cqdashboard.serializers import (
    DepartmentSerializer, ManagerDashboardSerializer, ManagerChartUploadSerializer, SplitAuditSerializer, TeamSerializer, TeamDashboardSerializer,
    HealthSystemSerializer, HealthSystemMembersSerializer, HealthSystemHospitalSerializer, HealthSystemTeamStatisticsSerializer, 
    HealthSystemHospitalDepartmentSerializer, HealthSystemHospitalProvidersSerializer, HospitalAccountsSerializer, InsuranceSerializer, EhrSerializer,
    ProviderSerializer, HospitalDepartmentDropDownSerializer, HealthSystemValidationSerializer, HospitalValidationSerializer, DepartmentValidationSerializer,
    MemberListSerializer
)

from cqclient.models import Client, Department, HealthSystem, Hospital, Insurance, Ehr, ProviderStatistics
from codequick.utils.mixins import SerializerClassMixin
from cqusers.serializers import SpecialtySerializer, SpecialtyAuditorSerializer

from cqdashboard.tasks import docx_to_pdf
from cqdashboard.utils import (
    health_system_inactive_spoc_user, hospital_inactive_spoc_user,
    department_inactive_spoc_user
)

from cqusers.serializers import LoginSerializer, ClientLoginSerializer,  SpecialtySerializer, SpecialtyAuditorSerializer
from cqusers.pagination import CustomPagination
from cqusers.models import CqUser, Specialty, CqTeam
from cqusers.tasks import send_email
from cqusers.utils import create_notification

from cqclient.models import Client, Department, HealthSystem, Hospital, Insurance, Ehr
from cqclient.utils import (
    all_specialties,
    get_health_system_clients,
    get_hospital_clients,
    get_department_clients,
    get_providers_clients,
)

default_storage = get_storage_class()()



class ManagerDashboard(SerializerClassMixin, viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ManagerDashboardSerializer
    serializer_action_classes = {
        # 'clients': LoginSerializer,
        'clients': ClientLoginSerializer,
    }
    pagination_class = CustomPagination
    parser_classes = [MultiPartParser, FormParser]

    filter_backends = [DjangoFilterBackend]
    # filterset_fields = ['status']
    queryset = Chart.objects.filter(is_deleted=False)


    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset().filter(is_split=False)).order_by('-urgent_flag', '-upload_date',)
        own = request.query_params.get('user', None)
        search = request.query_params.get('search', None)
        chart_status = request.query_params.get('status', None)
        urgent = request.query_params.get('urgent', None)

        if own:
            if request.user.role == 'MANAGER' or request.user.role == "AUDITOR":
                queryset = queryset.filter(auditor=request.user.id)

            elif request.user.role == 'QA':
                queryset = queryset.filter(qa=request.user.id)

        if team_id := self.request.query_params.get('team_id'):
            team_user = CqTeam.objects.get(id=team_id).members.filter(is_active=True, is_deleted=False).values_list("id", flat=True)
            queryset = queryset.filter(Q(auditor__in=team_user) | Q(qa__in=team_user))

        if user_id := self.request.query_params.get('user_id'):
            queryset = queryset.filter(Q(auditor=user_id) | Q(qa=user_id) | Q(client=user_id))

        if health_system_id:= self.request.query_params.get('health_system_id'):
            client_user_ids = get_health_system_clients(health_system_id)
            if client_user_ids == None:
                return Response({"message": "No HealthSystem matches the given query."}, status=status.HTTP_400_BAD_REQUEST)

            queryset = Chart.objects.filter(Q(client__user_id__in=client_user_ids) & ~Q(batch_id=None))

        if hospital_id:= self.request.query_params.get('hospital_id'):
            client_user_ids = get_hospital_clients(hospital_id)

            if not client_user_ids:
                return Response({"message": "No Hospital matches the given query."}, status=status.HTTP_400_BAD_REQUEST)

            queryset = Chart.objects.filter(Q(client__user_id__in=client_user_ids) & ~Q(batch_id=None))

        if department_id:=self.request.query_params.get('department_id'):
            client_user_ids = get_department_clients(department_id)

            if not client_user_ids:
                return Response({"message": "No Department matches the given query."}, status=status.HTTP_400_BAD_REQUEST)

            queryset = Chart.objects.filter(Q(client__user_id__in=client_user_ids) & ~Q(batch_id=None))

        if provider_id:=self.request.query_params.get('provider_id'):
            client_user_ids = get_providers_clients(provider_id)

            if not client_user_ids:
                return Response({"message": "No Providers matches the given query."}, status=status.HTTP_400_BAD_REQUEST)
            queryset = Chart.objects.filter(Q(client__user_id__in=client_user_ids) & ~Q(batch_id=None))

        if chart_status:
            queryset = queryset.filter(status__in=[i.upper().strip() for i in chart_status.split(",")])
            if request.user.role == 'QA' and 'AWAITING REVIEW' in chart_status.split(","):
                queryset = queryset.filter(specialty__in = request.user.specialties.all())

        if urgent:
            queryset = queryset.filter(urgent_flag=True)

        if search:
            search_cquser = CqUser.objects.annotate(search=SearchVector('first_name', 'last_name')).filter(search=search).values('id')

            if own:
                if request.user.role in ['MANAGER', 'AUDITOR']:
                    if chart_status in ['REBUTTAL', 'URGENT']:
                        queryset = queryset.filter(Q (status__icontains=search) | Q (qa__in=search_cquser) | 
                                            Q (specialty__name__icontains=search) | Q(chart_id__icontains=search))
                    else:
                        queryset = queryset.filter(Q (status__icontains=search) | Q (specialty__name__icontains=search) | Q(chart_id__icontains=search))

                elif request.user.role == "QA":
                    queryset = queryset.filter(Q (status__icontains=search) | Q (auditor__in=search_cquser) | 
                                           Q (specialty__name__icontains=search) | Q(chart_id__icontains=search))
            else:
                search_value = [i.upper().strip() for i in search.split(",")] if len([i.upper().strip() for i in search.split(",")]) > 1 else ''.join([i.upper().strip() for i in search.split(",")])
                queryset = queryset.filter(Q (status__icontains=search_value) | Q (status__in=search_value) | Q (client__user__in=search_cquser) | 
                                            Q (auditor__in=search_cquser)| Q (specialty__name__icontains=search) | Q(chart_id__icontains=search))


        if 'ordering' in request.query_params:
            if request.query_params['ordering'] == "client":
                queryset = queryset.annotate(fullname=Concat('client__user__first_name', Value(' '), 'client__user__last_name')).order_by("fullname") 
            elif request.query_params['ordering'] == "-client":
                queryset = queryset.annotate(fullname=Concat('client__user__first_name', Value(' '),  'client__user__last_name')).order_by("-fullname")
            else:
                queryset = queryset.order_by(request.query_params['ordering'].replace("updated_date", "chart_updated_date"))

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response({'results': serializer.data}, status=status.HTTP_200_OK)

    @transaction.atomic
    # API to submit the charts
    def create(self, request, *args, **kwargs):
        file_ids = request.data['file_ids'].split(',')
        client = request.data.get('client', None)

        if client is None:
            return Response({"message": "Client cannot be None", "error_code": 3004}, status=status.HTTP_400_BAD_REQUEST)

        client_user = CqUser.objects.get(id=client).client

        if len(file_ids) > 20:
            return Response({"message": "Maximum 20 files only can be uploaded", "error_code": 3004}, status=status.HTTP_400_BAD_REQUEST)

        chart_upload_file_ids = [file_ids for file_ids in ChartUpload.objects.filter(id__in=file_ids)]
        # _batch_id = int(timezone.now().timestamp())
        _batch_id = int(timezone.now().timestamp())
        prefix = client_user.health_system

        for each_ in chart_upload_file_ids:
            chart, created = Chart.objects.get_or_create(batch_id=_batch_id, client=client_user, upload_chart=each_.upload_chart, offline_upload_flag=True, total_pages=each_.total_pages)
            chart.chart_id = f"{prefix}{chart.id}"
            chart.save()
        create_notification(f"@ uploaded {len(file_ids)} new charts", client_user.user, CqUser.objects.filter(role='MANAGER', is_active=True, is_deleted=False), 'CHART UPLOADED', None)
        return Response(status=status.HTTP_201_CREATED)

    #Partial Update for three dot menu and total_pages, specialty, auditor
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = Chart.objects.get(id=kwargs['pk'])
        total_pages = request.data.get('total_page', None)
        specialty = request.data.get('assigned_specialties', None)
        auditor = request.data.get('assigned_auditor', None)
        urgent = request.data.get('urgent', None)
        request_missing_info = request.data.get('request_missing_info', None)

        if total_pages:
            if '-' in total_pages and '-' in instance.total_pages:
                if int(total_pages.split('-')[-1]) > int(instance.parent_chart.total_pages):
                    return Response({"message": "To page should not exceed the main chart total pages", "error_code": 4063}, status=status.HTTP_400_BAD_REQUEST)

            elif '-' in total_pages:
                if int(total_pages.split('-')[-1]) > int(instance.parent_chart.total_pages):
                    return Response({"message": "To page should not exceed the main chart total pages", "error_code": 4063}, status=status.HTTP_400_BAD_REQUEST)

            elif '-' in instance.total_pages:
                if int(total_pages) > int(instance.parent_chart.total_pages):
                    return Response({"message": "To page should not exceed the main chart total pages", "error_code": 4063}, status=status.HTTP_400_BAD_REQUEST)

                
            elif int(total_pages) > int(instance.total_pages):
                return Response({"message": "To page should not exceed the main chart total pages", "error_code": 4064}, status=status.HTTP_400_BAD_REQUEST)

            instance.total_pages = total_pages

        if specialty:
            instance.specialty = Specialty.objects.get(id=specialty)

        if auditor:
            if auditor == 'null':
                instance.auditor = None
                instance.status = 'AWAITING ASSIGNMENT'
            else:
                try:
                    user = get_object_or_404(CqUser, id=auditor)
                except Exception as e:
                    return Response({"message": str(e)}, status=status.HTTP_400_BAD_REQUEST)
                if user.role not in ['AUDITOR', 'MANAGER']:
                    return Response({"error": "Please choose an auditor."}, status=status.HTTP_400_BAD_REQUEST)
                instance.auditor = CqUser.objects.get(id=auditor)
                if instance.status == 'IN PROGRESS':
                    chart_history, created = ChartHistory.objects.get_or_create(chart=instance, user_type='AUDITOR')
                    chart_history.user = instance.auditor
                    chart_history.save()
                instance.status = 'AWAITING AUDIT'


        if urgent:
            instance.urgent_flag = True if urgent.lower() == 'true' else False

        if request_missing_info:
            instance.status = 'ON HOLD'
            html_body = """
            <h1 style="color:#023b93;">Welcome %s %s!</h1>
            <p style="font-size:15px;"> %s</p>
            """ %(
                request.user.first_name.title(),
                request.user.last_name.title(),
                request_missing_info
            )

            plain_body = """
            <h1 style="color:#023b93;">Welcome %s %s!</h1>
            <p style="font-size:15px;">%s</p>
            <p style="font-size:15px;"> This message is regarding missing information of the requested chart.</p>
            """ %(
                request.user.first_name.title(),
                request.user.last_name.title(),
                request_missing_info
            )

            _chart_id = instance.parent_chart.chart_id if instance.parent_chart else instance.chart_id
            _batch_id = f"{instance.client.health_system}{instance.batch_id}"

            send_email(
                    subject=f"CHART ID {_batch_id}-{_chart_id} - Audit Hold - Request Missing Information",
                    htmlBody=html_body,
                    plainBody = plain_body,
                    to = [instance.client.user.email],
                    cc = [request.user.email,],
                    bcc=["sankavi.boopathy@buildingblocks.la", "vishnu.kumar@buildingblocks.la", "vishnu@mailinator.com"]
                )

        instance.save()

        if instance.auditor == request.user:
            message = f"{instance.chart_id} has been assigned to self"
        else:
            message = f"{instance.chart_id} has been assigned to you"

        create_notification(message, request.user, [instance.auditor], 'CHART ASSIGNED', instance)    

        serializer = self.serializer_class(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        # serializer = UpdateManagerChartSerializer(instance)

        return Response({"result": serializer.data}, status=status.HTTP_200_OK)

    #Partial Delete offline_upload_flag=True
    def destroy(self, request, *args, **kwargs):
        instance = Chart.objects.get(id=kwargs['pk'])
        instance.is_deleted = True
        instance.save()
        # file_name = instance.upload_chart.name
        # folder_path = settings.MEDIAFILES_LOCATION
        # instance.upload_chart.delete()
        # default_storage.delete(folder_path)

        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], permission_classes=[permissions.IsAuthenticated], url_path='clients')
    def clients(self, request):
        queryset = CqUser.objects.filter(role='CLIENT', is_active=True, client__can_upload=True, client__is_approved=True, client__is_deleted=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


#Chart Upload
class ManagerChartUploadViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = ManagerChartUploadSerializer
    queryset = ChartUpload.objects.all().order_by('-uploaded_date')
    parser_classes = [MultiPartParser, FormParser]


    def create(self, request, *args, **kwargs): 
        client = request.data.get('client', None)
        chart_ = request.FILES['upload_chart']

        errors = {}
        if client is None:
            errors.update({"message": "Client cannot be None", "error_code": 3004})

        client_user = CqUser.objects.get(id=client).client
        request.data['client'] = client_user.id


        if client_user.is_primary == True:
            errors.update({"message": "Client is a primary account owner and can upload only compliance charts.", "error_code": 3005})

        if client_user.can_upload == False:
            errors.update({"message": "Client does not have the permission to upload chart.", "error_code": 3006})

        if chart_.name.split('.')[-1] not in ['pdf', 'hl7', 'doc', 'docx', 'xls', 'xlsx']:
            errors.update({"message": "Wrong file format", "error_code": 3001})

        elif chart_.size <= 0:
            errors.update({"message": "File is empty", "error_code": 3002})

        elif round(chart_.size / (1024 * 1024), 2) > 300:
            errors.update({"message": "File size exceeds 300mb", "error_code": 3003})

        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        # To get the total no of pages of uploaded pdf
        if chart_.name.split('.')[-1] == "pdf":
            readpdf = PdfFileReader(chart_)
            totalpages = readpdf.numPages

        # To get the total no of pages for doc and docx
        elif chart_.name.split('.')[-1] in ("doc", "docx"):
            # To convert doc to pdf
            doc_obj = docx_to_pdf(chart_)
            inputpdf = PdfFileReader(doc_obj)
            totalpages = inputpdf.numPages

        # To get total sheets for XLS and XLSX:
        elif chart_.name.split('.')[-1] in ('xls' ,'xlsx'):
            xl = pd.ExcelFile(chart_)
            totalpages = len(xl.sheet_names)


        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(total_pages=totalpages)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    # def destroy(self, request, *args, **kwargs):
    #     instance = self.get_object()
    #     file_name = instance.upload_chart.name
    #     folder_path = settings.MEDIAFILES_LOCATION

    #     instance.upload_chart.delete()
    #     default_storage.delete(folder_path)

    #     return Response(status=status.HTTP_400_BAD_REQUEST)


class SplitAuditViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = SplitAuditSerializer
    queryset = Chart.objects.all()

    @action(detail=True, methods=['post'], permission_classes=[permissions.AllowAny, ])
    def post(self, request, pk):
        chart = Chart.objects.get(id=pk)

        # Page no partial Validation:
        for page_nos in request.data:
            if page_nos['page_number_from'] < 1:
                return Response({"message": "Please check the page no for from page", "error_code": 4060}, status=status.HTTP_400_BAD_REQUEST)

        from_page = [from_page['page_number_from'] for from_page in request.data]
        to_page = [to_page['page_number_to'] for to_page in request.data]

        # for (from_page, to_page) in zip(from_page[1:], to_page):
        #     if from_page <= to_page:
        #         return Response({"message": "From page should not less than or equal to previous split to_page", "error_code": 4061}, status=status.HTTP_400_BAD_REQUEST)

        split_chart_pages = [page_nos['page_number_to'] for page_nos in request.data]
        if chart.upload_chart.name.split('.')[-1] in ("xlsx", "xls"):
            wb = load_workbook(filename = chart.upload_chart)
            if max(split_chart_pages) > len(wb.worksheets):
                return Response({"message": "To page should not exceed the main chart total pages", "error_code": 4062}, status=status.HTTP_400_BAD_REQUEST)

        elif max(split_chart_pages) > int(chart.total_pages):
            return Response({"message": "To page should not exceed the main chart total pages", "error_code": 4062}, status=status.HTTP_400_BAD_REQUEST)

        # if sum(split_chart_pages) > int(chart.total_pages):
        #     return Response({"message": "Total pages of split audit charts should not exceed the main chart total pages", "error_code": 4063}, status=status.HTTP_400_BAD_REQUEST)

        if chart.upload_chart.name.split('.')[-1] in ("doc", "docx"):
            #Convert Doc or Docx to PDF
            doc_obj = docx_to_pdf(chart.upload_chart)
            inputpdf = PdfFileReader(doc_obj)
            total_pages = inputpdf.numPages

        for each_ in request.data:

            if Chart.objects.filter(chart_id=f"{chart.chart_id}-{each_['chart_id']}").exists():
                return Response({"message": "Chart Name already exists", "error_code": 4064}, status=status.HTTP_400_BAD_REQUEST)

            if Chart.objects.filter(chart_id=each_['chart_id']).exists():
                return Response({"message": "Chart Name already exists", "error_code": 4065}, status=status.HTTP_400_BAD_REQUEST)

            if chart.upload_chart.name.split('.')[-1] == "pdf":
                inputpdf = PdfFileReader(chart.upload_chart)
                total_pages = inputpdf.numPages

                with PdfFileWriter() as output:
                    for page_numbers in range(inputpdf.numPages)[each_['page_number_from']-1 : each_['page_number_to']]:

                        file_name = f"{each_['chart_id']}.pdf"
                        output.addPage(inputpdf.getPage(page_numbers))  
                        with open(file_name, "ab") as outputStream:
                            output.write(outputStream)
                        splitted_chart = File(open(file_name, mode='rb'), name=file_name)
                        split_chart_total_pages = PdfFileReader(splitted_chart).numPages


            if chart.upload_chart.name.split('.')[-1] in ("doc", "docx"):

                total_pages = inputpdf.numPages

                with PdfFileWriter() as output:
                    for page_numbers in range(inputpdf.numPages)[each_['page_number_from']-1 : each_['page_number_to']]:
                        file_name = f"{each_['chart_id']}.pdf"
                        output.addPage(inputpdf.getPage(page_numbers))  
                        with open(file_name, "ab") as outputStream:
                            output.write(outputStream)
                        splitted_pdf = File(open(file_name, mode='rb'), name=file_name)
                        split_chart_total_pages = PdfFileReader(splitted_pdf).numPages

                # Convert PDF to Doc or Docx:
                converted_doc = Converter(splitted_pdf)
                converted_doc.convert(f"{each_['chart_id']}.{chart.upload_chart.name.split('.')[-1]}")
                converted_doc.close
                splitted_chart = File(open(f"{each_['chart_id']}.{chart.upload_chart.name.split('.')[-1]}", mode='rb'), name=f"{each_['chart_id']}.{chart.upload_chart.name.split('.')[-1]}")
                os.remove(splitted_pdf.name)


            if chart.upload_chart.name.split('.')[-1] in ("xls", "xlsx"):
                wb = load_workbook(filename = chart.upload_chart)
                ws = wb.worksheets[:each_['page_number_to']]

                for sheets_ in ws:
                    if len(ws) > 1:
                        wb.remove_sheet(sheets_)
                    wb.save(f"{each_['chart_id']}.{chart.upload_chart.name.split('.')[-1]}")
                    excel_file = glob.glob("*.xlsx") + glob.glob("*.xls")
                    for file_ in excel_file:
                        splitted_chart = File(open(file_, mode='rb'), name=file_)

            total_page = f"{each_['page_number_from']}-{each_['page_number_to']}"
            specialty = Specialty.objects.get(id=each_['specialty'])
            _user = CqUser.objects.get(id=each_['auditor'])

            if _user.role not in ['AUDITOR', 'MANAGER']:
                return Response({"error": "Please choose an auditor."}, status=status.HTTP_400_BAD_REQUEST)

            split_chart = Chart.objects.create(chart_id=each_['chart_id'], client=chart.client, specialty=specialty, auditor=_user, status="AWAITING AUDIT", upload_chart=splitted_chart, parent_chart=chart, total_pages=total_page)
            os.remove(splitted_chart.name)
            chart.is_split = True
            chart.save()
        return Response(status=status.HTTP_201_CREATED)


class TeamViewSet(SerializerClassMixin, viewsets.ModelViewSet):
    queryset = CqTeam.objects.filter(is_active=True).order_by('-id')
    serializer_class = TeamSerializer
    serializer_action_classes ={
        'list': TeamDashboardSerializer,
        'members_list': MemberListSerializer
    }
    permission_classes = (permissions.IsAuthenticated,)
    pagination_class = CustomPagination

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        if search := self.request.query_params.get('search'):
            search_cquser = CqUser.objects.annotate(search=SearchVector('first_name', 'last_name')).filter(is_active=True, search=search).values('id')
            queryset = queryset.filter(Q(name__icontains=search) | Q(members__in=search_cquser)).distinct()

        if ordering := request.query_params.get('ordering'):
            queryset = queryset.order_by(ordering)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer_class()(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def create(self, request, *args, **kwargs):
        name = request.data.get('name')
        members = request.data.get('members')
        specialties = request.data.get('specialties')
        errors = {}
        if CqTeam.objects.filter(name__iexact=name).exists():
            errors.update({'name': 'Team with this name already exist'})
        if name.strip() == "":
            errors.update({'name': 'Team name can not be empty'})
        if members == list():
            errors.update({'members': 'Members field can not be empty'})
        if specialties == list():
            errors.update({'specialties': 'Specialties field can not be empty'})

        if errors:
            response = {'error_code': 4080, 'errors': errors, 'message':[errors[key] for key in errors]}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

        request.data['name'] = name.strip()
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.is_active:
            serializer = TeamDashboardSerializer(instance)    
            return Response(serializer.data)
        return Response({"message": "Requested team is deleted"})

    def update(self, request, *args, **kwargs):
        name = request.data.get('name', '').strip()
        members = request.data.get('members')
        specialties = request.data.get('specialties')
        errors = {}
        # if CqTeam.objects.filter(lower_name=name.lower()).exists():
        #     errors.update({'name': 'Team with this name already exist'})
        if members == list():
            errors.update({'members': 'Members field can not be empty'})

        if specialties == list():
            errors.update({'specialties': 'Specialties field can not be empty'})

        if name:
            if CqTeam.objects.filter(name__iexact=name).exclude(pk=kwargs['pk']).exists():
                errors.update({'name': 'Another team with this name already exist'})
        else:
            errors.update({'name': 'Team name cannot be empty'})

        if errors:
            response = {'error_code': 4080, 'errors': errors, 'message':[errors[key] for key in errors]}
            return Response(response, status=status.HTTP_400_BAD_REQUEST)

        request.data['name'] = name.strip()
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response({"message": "Team updated successfully"})

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_active = False
        instance.name = instance.name+f"_deleted_{instance.id}"
        instance.save()
        return Response({'message': "Team deleted successfully"}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], url_path='members_list')
    def members_list(self, request):
        specialties_ids = [int(specialty) for specialty in self.request.query_params.get('specialties').split(',')]
        queryset = CqUser.objects.filter(role__in=['AUDITOR', 'QA'], is_active=True, is_deleted=False, specialties__id__in=specialties_ids).order_by('role')
        serializer = self.get_serializer(queryset, many=True, context={'specialties_ids': specialties_ids})
        return Response(serializer.data)


class SpecialtyViewSet(SerializerClassMixin, viewsets.ModelViewSet):
    permission_classes = (permissions.AllowAny,)
    serializer_class = SpecialtySerializer
    serializer_action_classes = {
        "auditor":SpecialtyAuditorSerializer,
    }
    queryset = Specialty.objects.all()

    def list(self, request, *args, **kargs):
        search = self.request.query_params.get('search', '').strip()
        # hospital_id = self.request.query_params.get('hospital_id')

        if hospital_id:=request.query_params.get('hospital_id'):
            # Get Specialty Ids Assigend to Hospital Pk
            specialty_ids = [ each_.specialty.id for each_ in Department.objects.filter(hospital=hospital_id, is_active=True, is_deleted=False) ]
            queryset = Specialty.objects.filter(id__in=specialty_ids).order_by('name').values('id', 'name')

        else:
            queryset = Specialty.objects.filter(type='MANAGER').order_by('name').values('id', 'name')

        if search:
            queryset = Specialty.objects.filter(name__istartswith=search).order_by('name').values('id', 'name')
            if not queryset:
                queryset = list(filter(lambda each: search.casefold() in each['name'].casefold(), all_specialties))

        return Response(queryset, status=status.HTTP_200_OK)


    @action(detail=False, methods=['get'], url_path='auditors')
    def auditor(self,request):
        queryset = self.get_queryset()
        if search:=request.query_params.get('search'):
            search_cquser = CqUser.objects.annotate(search=SearchVector('first_name', 'last_name')).filter(is_active=True, search=search).values('id')
            queryset = queryset.filter(Q(name__icontains=search) | Q(cqusers__first_name__icontains=search) | Q(cqusers__in=search_cquser))

        serializer = self.get_serializer_class()(queryset, many=True)
        return Response(serializer.data)


class HealthSystemViewSet(SerializerClassMixin, viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = HealthSystemSerializer
    serializer_action_classes = {
        "clients": HealthSystemMembersSerializer,
        "hospitals": HealthSystemHospitalSerializer,
        "department": HealthSystemHospitalDepartmentSerializer,
        "providers": HealthSystemHospitalProvidersSerializer,
        "team_statistics": HealthSystemTeamStatisticsSerializer,
        "validation": HealthSystemValidationSerializer,
    }
    pagination_class = CustomPagination
    queryset = HealthSystem.objects.filter(is_deleted=False)
    # ordering_fields = ('active_audits',)

    def list(self, request, *args, **kargs):
        queryset = self.get_queryset().order_by('-id')

        if search:=request.query_params.get('search'):
            queryset = queryset.filter(Q(name__icontains=search) | Q(hospital_health_system__specialty__name__icontains=search)).distinct()

        if ordering:=request.query_params.get('ordering'):
            queryset = queryset.order_by(ordering)

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @method_decorator(never_cache)
    @transaction.atomic
    def create(self, request, *args, **kwargs):

        name = request.data.get('name', '').strip()
        prefix = request.data.get('prefix', '').strip()
        address = request.data.get('address', '').strip()
        physicians_group = request.data.get('physicians_group', None)

        specialties = request.data.get("specialties", None)
        insurance = request.data.get("insurance", None)
        ehr = request.data.get("ehr", None)

        # Validations:
        errors = {}

        if request.data.get('account_contact'):
            for index, data_ in enumerate(request.data['account_contact']):
                errors[index] = {}
                first_name = data_.get("first_name", '').strip()
                last_name = data_.get("last_name", '').strip()
                email = data_.get("email", '').strip()


                if first_name == "":
                    errors[index].update({'first_name': 'Firstname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', first_name):
                    errors[index].update({'first_name': 'Firstname can contain only alphabets'})



                if last_name == "":
                    errors[index].update({'last_name': 'Lastname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', last_name):
                    errors[index].update({'last_name': 'Lastname can contain only alphabets'})


                if email:
                    if CqUser.objects.filter(email__iexact=email).exists():
                        errors[index].update({'email': f'Email {email} already taken'})

                else:
                    errors[index].update({'email': f'Email cannot be empty'})

                if not errors[index]:
                    errors.clear()

        if prefix:
            if HealthSystem.objects.filter(prefix__iexact=prefix).exists():
                errors.update({'prefix': 'This prefix has already been taken.'})

        else:
            errors.update({'prefix': 'Prefix cannot be empty.'})

        if name:
            if not re.search('^[A-Za-z ]+$', name):
                errors.update({'name': 'Given Health System can contain only alphabets!'})

            if HealthSystem.objects.filter(name__iexact=name).exists():
                errors.update({'name': 'Given Health System name already exists!'})

        else:
            errors.update({'name': 'Health System name cannot be empty.'})

        if specialties:
            if specialties['exist_ids']:
                for id_ in specialties['exist_ids']:
                    if not Specialty.objects.filter(id=id_).exists():
                        errors.update({"specialty": f"Given Specialty {id_} is not present."})

        if insurance:
            if insurance['exist_ids']:
                for id_ in insurance['exist_ids']:
                    if not Insurance.objects.filter(id=id_).exists():
                        errors.update({"insurance": f"Given Insurance {id_} is not present."})

        if ehr:
            if ehr['exist_ids']:
                for id_ in ehr['exist_ids']:
                    if not Ehr.objects.filter(id=id_).exists():
                        errors.update({"ehr": f"Given Ehr {id_} is not present."})

        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        health_system = HealthSystem.objects.create(name=name, prefix=prefix, type='PHYSGRP', address=address) if physicians_group else HealthSystem.objects.create(name=name, prefix=prefix, type='HS', address=address)
        address_ = address if address else None

        if physicians_group:
            hospital = Hospital.objects.create(name=name, address=address_, health_system=health_system, patients_per_month=0)

        if request.data.get('account_contact'):
            for data_ in request.data['account_contact']:
                user = CqUser.objects.create(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'], role="CLIENT")
                client_user = Client.objects.create(user=user, user_type="PHYSICIANS GROUP", is_primary=data_['is_primary']) if physicians_group else Client.objects.create(user=user, user_type="HEALTH SYSTEM", is_primary=data_['is_primary'])

                health_system.spoc.add(client_user)

                if physicians_group:
                    hospital.spoc.add(client_user)

                if specialties:
                    if specialties['exist_ids']:
                        # [user.specialties.add(each_) for each_ in specialties['exist_ids']]

                        if physicians_group:
                            [hospital.specialty.add(each_) for each_ in specialties['exist_ids']]

                    if specialties['new_spec_name']:
                        for new_spec_ in specialties['new_spec_name']:
                            data, new_spec = Specialty.objects.get_or_create(name=new_spec_)
                            # user.specialties.add(data)

                            if physicians_group:
                                hospital.specialty.add(data)

                uid = urlsafe_base64_encode(force_bytes(user.id))
                token = PasswordResetTokenGenerator().make_token(user)

                if request.headers['Origin'] in ['https://dev-manager.codequick.com', 'https://dev-client.codequick.com',]:
                    front_end_url = {
                        "client": "https://dev-client.codequick.com",
                    }

                elif request.headers['Origin'] in ['https://qa-manager.codequick.com', 'https://qa-qa.codequick.com', 'https://qa-client.codequick.com', 'https://qa-auditor.codequick.com']:
                    front_end_url = {
                        "manager": "https://qa-manager.codequick.com",
                        "qa": "https://qa-qa.codequick.com",
                        "client": "https://qa-client.codequick.com",
                        "auditor": "https://qa-auditor.codequick.com",
                    }

                elif request.headers['Origin'] in ['https://staging-manager.codequick.com', 'https://staging-qa.codequick.com', 'https://staging-client.codequick.com', 'https://staging-auditor.codequick.com']:
                    front_end_url = {
                        "manager": "https://staging-manager.codequick.com",
                        "qa": "https://staging-qa.codequick.com",
                        "client": "https://staging-client.codequick.com",
                        "auditor": "https://staging-auditor.codequick.com",
                    }

                else:
                    front_end_url = {
                        "client": "http://localhost:3000"
                    }

                url = f"{front_end_url['client']}/auth/reset-password?uid={uid}&token={token}&email={user.email}/"
                reset_url = "<a href =" + url + "> Register </a>"

                html_body = """
                    <b><h1 style="color:#023b93;">Welcome %s %s!</h1>
                    </b>
                    <p style="font-size:15px;">Your account has been created successfully %s.</p>
                    <p style="font-size:15px;">We request you to click on the link below to reset your password and get started</p>
                        %s<br>
                        <br>
                    """ %(
                        user.first_name.title(),
                        user.last_name.title(),
                        user.email,
                        reset_url,
                    )

                plain_body = """
                    <b>
                    <h1 style="color:#023b93;">Welcome %s %s!</h1>
                    <br><br>
                    <p>Your account has been created successfully %s.</p>
                        We request you to click on the link below to reset your password and get started<br>
                        %s<br>
                        <br>
                    """ %(
                        user.first_name.title(),
                        user.last_name.title(),
                        user.email,
                        reset_url
                    )

                send_email(
                    subject="Welcome to CodeQuick",
                    htmlBody=html_body,
                    plainBody = plain_body,
                    to=[user.email,],
                    bcc=["vishnu.kumar@buildingblocks.la", "sankavi.boopathy@buildingblocks.la", "vishnu@mailinator.com"]
                )

        if insurance:
            if insurance['exist_ids']:
                [health_system.insurance.add(ins_id) for ins_id in insurance['exist_ids']]

                if physicians_group:
                    [hospital.insurance.add(ins_id) for ins_id in insurance['exist_ids']]
 
            if insurance['new_ins_name']:
                for new_ins_ in insurance['new_ins_name']:
                    data, new_ins = Insurance.objects.get_or_create(name=new_ins_)
                    health_system.insurance.add(data)

                    if physicians_group:
                        hospital.insurance.add(data)

        if ehr:
            if ehr['exist_ids']:
                [health_system.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]

                if physicians_group:
                    [hospital.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]
 
            if ehr['new_ehr_name']:
                for new_ehr_ in ehr['new_ehr_name']:
                    data, new_ehr = Ehr.objects.get_or_create(name=new_ehr_)
                    health_system.ehr.add(data)

                    if physicians_group:
                        hospital.ehr.add(data)

        serializer = self.get_serializer(health_system)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = HealthSystem.objects.get(id=kwargs['pk'])

        name = request.data.get('name', '').strip()
        prefix = request.data.get('prefix', '').strip()
        address = request.data.get('address', '').strip()
        physicians_group = request.data.get('physicians_group', None)

        #specialties = request.data.get("specialties", None)
        insurance = request.data.get("insurance", None)
        ehr = request.data.get("ehr", None)

        # Validation:
        errors = {}

        if request.data.get('account_contact'):
            for index, data_ in enumerate(request.data['account_contact']):
                errors[index] = {}
                first_name = data_.get("first_name", '').strip()
                last_name = data_.get("last_name", '').strip()

                if data_.get("id") == None:
                    email = data_.get("email", '').strip()

                    if email == "":
                        errors[index].update({'email': f"Email cannot be blank"})

                    if CqUser.objects.filter(email__iexact=email).exists():
                        errors[index].update({'email': f"Email {data_['email']} already taken"})

                    if first_name == "":
                        errors[index].update({'first_name': 'Firstname cannot be blank'})

                    elif not re.search('^[A-Za-z ]+$', first_name):
                        errors[index].update({'first_name': 'Firstname can contain only alphabets'})

                    if last_name == "":
                        errors[index].update({'last_name': 'Lastname cannot be blank'})

                    elif not re.search('^[A-Za-z ]+$', last_name):
                        errors[index].update({'last_name': 'Lastname can contain only alphabets'})


                if not errors[index]:
                    errors.clear()

        # if specialties:
        #     if specialties['exist_ids']:
        #         for id_ in specialties['exist_ids']:
        #             if not Specialty.objects.filter(id=id_).exists():
        #                 errors.update({"specialty": f"Given Specialty {id_} is not present."})

        if insurance:
            if insurance['exist_ids']:
                for id_ in insurance['exist_ids']:
                    if not Insurance.objects.filter(id=id_).exists():
                        errors.update({"insurance": f"Given Insurance {id_} is not present."})

        if ehr:
            if ehr['exist_ids']:
                for id_ in ehr['exist_ids']:
                    if not Ehr.objects.filter(id=id_).exists():
                        errors.update({"ehr": f"Given Ehr {id_} is not present."})

        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if HealthSystem.objects.filter(name__iexact=name).exclude(name=name).exists():
                errors.update({'name': 'Health System with this name already exists.'})
            else:
                HealthSystem.objects.filter(id=instance.id).update(name=name)

        except Exception as e:
            return Response({"name": "Health System with this name already exists."}, status=status.HTTP_400_BAD_REQUEST)

        address_ = address if address else None
        Hospital.objects.filter(health_system=instance).update(address=address_)
        hospital_ = Hospital.objects.filter(health_system=instance)

        if request.data.get('account_contact'):
            for data_ in request.data['account_contact']:
                if data_.get("id") == None:
                    user = CqUser.objects.create(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'], role="CLIENT")
                    client_user = Client.objects.create(user=user, user_type="HEALTH SYSTEM", is_primary=data_['is_primary'])
                    instance.spoc.add(client_user)

                    if instance.type == 'PHYSGRP':
                        for hos_ in hospital_:
                            hos_.spoc.add(client_user)

                    uid = urlsafe_base64_encode(force_bytes(user.id))
                    token = PasswordResetTokenGenerator().make_token(user)

                    if request.headers['Origin'] in ['https://dev-manager.codequick.com', 'https://dev-client.codequick.com',]:
                        front_end_url = {
                            "client": "https://dev-client.codequick.com",
                        }
                
                    elif request.headers['Origin'] in ['https://qa-manager.codequick.com', 'https://qa-qa.codequick.com', 'https://qa-client.codequick.com', 'https://qa-auditor.codequick.com']:
                        front_end_url = {
                            "manager": "https://qa-manager.codequick.com",
                            "qa": "https://qa-qa.codequick.com",
                            "client": "https://qa-client.codequick.com",
                            "auditor": "https://qa-auditor.codequick.com",
                        }
                    elif request.headers['Origin'] in ['https://staging-manager.codequick.com', 'https://staging-qa.codequick.com', 'https://staging-client.codequick.com', 'https://staging-auditor.codequick.com']:
                        front_end_url = {
                            "manager": "https://staging-manager.codequick.com",
                            "qa": "https://staging-qa.codequick.com",
                            "client": "https://staging-client.codequick.com",
                            "auditor": "https://staging-auditor.codequick.com",
                        }
                    else:
                        front_end_url = {
                            "client": "http://localhost:3000"
                        }

                    url = f"{front_end_url['client']}/auth/reset-password?uid={uid}&token={token}&email={user.email}/"
                    reset_url = "<a href =" + url + "> Register </a>"

                    html_body = """
                        <b><h1 style="color:#023b93;">Welcome %s %s!</h1>
                        </b>
                        <p style="font-size:15px;">Your account has been created successfully %s.</p>
                        <p style="font-size:15px;">We request you to click on the link below to reset your password and get started</p>
                            %s<br>
                            <br>
                        """ %(
                            user.first_name.title(),
                            user.last_name.title(),
                            user.email,
                            reset_url,
                        )

                    plain_body = """
                        <b>
                        <h1 style="color:#023b93;">Welcome %s %s!</h1>
                        <br><br>
                        <p>Your account has been created successfully %s.</p>
                            We request you to click on the link below to reset your password and get started<br>
                            %s<br>
                            <br>
                        """ %(
                            user.first_name.title(),
                            user.last_name.title(),
                            user.email,
                            reset_url
                        )

                    send_email(
                        subject="Welcome to CodeQuick",
                        htmlBody=html_body,
                        plainBody = plain_body,
                        to=[user.email,],
                        bcc=["vishnu.kumar@buildingblocks.la", "sankavi.boopathy@buildingblocks.la", "vishnu@mailinator.com"]
                    )

                else:
                    CqUser.objects.filter(id=data_['id']).update(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'])
                    Client.objects.filter(user__id=data_['id']).update(is_primary=data_['is_primary'])
                    user = CqUser.objects.get(id=data_['id'])


        if insurance:
            instance.insurance.clear()
            [hosp_.insurance.clear() for hosp_ in hospital_]

            if insurance['exist_ids']:
                [instance.insurance.add(ins_id) for ins_id in insurance['exist_ids']]
                for hosp_ in hospital_:
                    [hosp_.insurance.add(ins_id) for ins_id in insurance['exist_ids']]

            if insurance['new_ins_name']:
                for new_ins_ in insurance['new_ins_name']:
                    data, new_ins = Insurance.objects.get_or_create(name=new_ins_)
                    instance.insurance.add(data)
                    for hosp_ in hospital_:
                        hosp_.insurance.add(data)

        if ehr:
            instance.ehr.clear()
            [hosp_.ehr.clear() for hosp_ in hospital_]

            if ehr['exist_ids']:
                [instance.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]
                for hosp_ in hospital_:
                    [hosp_.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]

            if ehr['new_ehr_name']:
                for new_ehr_ in ehr['new_ehr_name']:
                    data, new_ehr = Ehr.objects.get_or_create(name=new_ehr_)
                    instance.ehr.add(data)
                    for hosp_ in hospital_:
                        hosp_.ehr.add(data)

        instance.name = name
        instance.prefix = prefix
        instance.save()

        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def destroy(self, request, pk):
        instance = self.get_object()
        instance.is_active = False
        instance.is_deleted = True

        instance.save()
        return Response({"message":"HealthSystem deleted successfully"})

    @action(detail=True, methods=['get',], url_path="clients")
    def clients(self, request, *args, **kwargs):

        if hospital_id := request.query_params.get('hospital_id'):
            client_ids = get_hospital_clients(request.query_params['hospital_id'])
            if not client_ids:
                return Response({"message": "No Hospital matches the given query."}, status=status.HTTP_400_BAD_REQUEST)

        elif department_id := request.query_params.get('department_id'):
            client_ids = get_department_clients(request.query_params['department_id'])
            if not client_ids:
                return Response({"message": "No Department matches the given query."}, status=status.HTTP_400_BAD_REQUEST)

        elif providers_id := request.query_params.get('providers_id'):
            client_ids = get_providers_clients(request.query_params['providers_id'])
            if not client_ids:
                return Response({"message": "No Providers matches the given query."}, status=status.HTTP_400_BAD_REQUEST)

        else:
            client_ids = get_health_system_clients(kwargs['pk'])
            if not client_ids:
                return Response({"message": "No HealthSystem matches the given query."}, status=status.HTTP_400_BAD_REQUEST)
    
        queryset = CqUser.objects.filter(id__in=list(client_ids), is_active=True, is_deleted=False)

        if search := self.request.query_params.get('search'):
            queryset = queryset.filter(Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(email__icontains=search))

        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get',], url_path="hospitals")
    def hospitals(self, request, pk=None):

        hospital_ids = [user['id'] for each_ in self.get_queryset().filter(id=pk) for user in each_.hospital_health_system.filter(is_active=True, is_deleted=False).values('id')]
        queryset = Hospital.objects.filter(id__in=hospital_ids, is_active=True, is_deleted=False)

        if search := self.request.query_params.get('search'):
            queryset = queryset.filter(name__icontains=search)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get',], url_path="department")
    def department(self, request, pk=None):
        department_ids = [department_['id'] for each_ in self.get_queryset().filter(id=pk) for hospital in each_.hospital_health_system.all() for department_ in hospital.department_hospital.values('id')]
        queryset = Department.objects.filter(id__in=department_ids, is_active=True, is_deleted=False)
        if search := self.request.query_params.get('search'):
            queryset = queryset.filter(specialty__name__icontains=search)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get',], url_path="providers")
    def providers(self, request, pk=None):
        providers_ids = [provider_['id'] for each_ in self.get_queryset().filter(id=pk) for hospital in each_.hospital_health_system.all() for provider_ in hospital.department_hospital.values('id')]
        queryset = Department.objects.filter(id__in=providers_ids, is_active=True, is_deleted=False)

        if search := self.request.query_params.get('search'):
            queryset = queryset.filter(specialty__name__icontains=search)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get',], url_path="team-statistics")
    def team_statistics(self, request, pk=None):

        if health_system_id := request.query_params.get('health_system_id'):
            queryset = self.get_queryset().filter(id=health_system_id)

        if hospital_id := request.query_params.get('hospital_id'):
            queryset = Hospital.objects.filter(id=hospital_id, is_active=True, is_deleted=False)

        if department_id := request.query_params.get('department_id'):
            queryset = Department.objects.filter(id=department_id, is_active=True, is_deleted=False)

        if provider_id := request.query_params.get('provider_id'):
            queryset = Department.objects.filter(id=provider_id, is_active=True, is_deleted=False)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get',], url_path="validation")
    def validation(self, request):
        queryset = HealthSystem.objects.filter(is_deleted=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class HospitalViewSet(SerializerClassMixin, viewsets.ModelViewSet):
    queryset = Hospital.objects.filter(is_deleted=False)
    serializer_class = HospitalAccountsSerializer
    permission_classes = (permissions.IsAuthenticated,)
    serializer_action_classes = {
        "validation": HospitalValidationSerializer,
    }

    def list(self, request):

        ordering = self.request.query_params.get('ordering') if self.request.query_params.get('ordering') else "id"

        if health_system :=self.request.query_params.get('health_system'):
            queryset = self.get_queryset().filter(health_system=health_system).order_by(ordering)
        else:
            queryset = self.get_queryset().order_by(ordering)

        if search := self.request.query_params.get('search'):
            # queryset = queryset.filter(name__istartswith=search)
            queryset = queryset.filter(name__icontains=search)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def retrieve(self, request, pk):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # Passing health system id in pk:
    @transaction.atomic
    @action(detail=True, methods=['post'])
    def post(self, request, pk):

        name = request.data.get("name", '').strip()
        address = request.data.get("address", '').strip()
        patients_per_month = request.data.get("patients_per_month", 0)
        specialties = request.data.get("specialties", None)
        insurance = request.data.get("insurance", None)
        ehr = request.data.get("ehr", None)
        health_system_name = request.data.get("health_system_name", '').strip()

        # Validations:
        errors = {}

        if request.data.get('account_contact'):
            for index, data_ in enumerate(request.data['account_contact']):
                errors[index] = {}
                first_name = data_.get("first_name", "").strip()
                last_name = data_.get("last_name", "").strip()
                email = data_.get("email", "").strip()


                if first_name == "":
                    errors[index].update({'first_name': 'Firstname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', first_name):
                    errors[index].update({'first_name': 'Firstname can contain only alphabets'})


                if last_name == "":
                    errors[index].update({'last_name': 'Lastname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', last_name):
                    errors[index].update({'last_name': 'Lastname can contain only alphabets'})

                if email == "":
                    errors[index].update({'email': f"Email cannot be blank"})

                if CqUser.objects.filter(email__iexact=email).exists():
                    errors[index].update({'email': f'Email {email} already taken'})


                if not errors[index]:
                    errors.clear()

        if name:

            if not re.search('^[A-Za-z ]+$', name):
                errors.update({'name': 'Hospital name can contain only alphabets'})

            if Hospital.objects.filter(name=name).exists():
                errors.update({'name': 'Given Hospital name already exists!'})

        else:
            errors.update({'name': 'Hospital name cannot be empty.'})

        if not HealthSystem.objects.filter(id=pk):
            errors.update({"health_system": "Given Health System is not present in the table."})
        else:
            health_system = HealthSystem.objects.get(id=pk)

        if health_system_name:
            if health_system_name != health_system.name:
                errors.update({"health_system_name": "Please give the proper health_system_name."})

        if specialties:
            if specialties['exist_ids']:
                for id_ in specialties['exist_ids']:
                    if not Specialty.objects.filter(id=id_).exists():
                        errors.update({"specialty": f"Given Specialty {id_} is not present."})

        if insurance:
            if insurance['exist_ids']:
                for id_ in insurance['exist_ids']:
                    if not Insurance.objects.filter(id=id_).exists():
                        errors.update({"insurance": f"Given Insurance {id_} is not present."})

        if ehr:
            if ehr['exist_ids']:
                for id_ in ehr['exist_ids']:
                    if not Ehr.objects.filter(id=id_).exists():
                        errors.update({"ehr": f"Given Ehr {id_} is not present."})

        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        address_ = address if address else None
        hospital_ = Hospital.objects.create(
            name=name,
            address=address_,
            patients_per_month=patients_per_month,
            health_system=health_system,
            goal = health_system.goal,
            top_parameter = health_system.top_parameter,
            bottom_parameter = health_system.bottom_parameter,
            is_active=False
            )

        if request.data.get('account_contact'):
            for data_ in request.data['account_contact']:
                user = CqUser.objects.create(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'], role="CLIENT")
                client_user = Client.objects.create(user=user, user_type="HOSPITAL", is_primary=data_['is_primary'])
                hospital_.spoc.add(client_user)

                if specialties:
                    if specialties['exist_ids']:
                        # [user.specialties.add(each_) for each_ in specialties['exist_ids']]
                        [hospital_.specialty.add(each_) for each_ in specialties['exist_ids']]

                    if specialties['new_spec_name']:
                        for new_spec_ in specialties['new_spec_name']:
                            data, new_spec = Specialty.objects.get_or_create(name=new_spec_)
                            # user.specialties.add(data)
                            hospital_.specialty.add(data)

                uid = urlsafe_base64_encode(force_bytes(user.id))
                token = PasswordResetTokenGenerator().make_token(user)

                if request.headers['Origin'] in ['https://dev-manager.codequick.com', 'https://dev-client.codequick.com',]:
                        front_end_url = {
                            "client": "https://dev-client.codequick.com",
                        }

                elif request.headers['Origin'] in ['https://qa-manager.codequick.com', 'https://qa-qa.codequick.com', 'https://qa-client.codequick.com', 'https://qa-auditor.codequick.com']:
                    front_end_url = {
                        "manager": "https://qa-manager.codequick.com",
                        "qa": "https://qa-qa.codequick.com",
                        "client": "https://qa-client.codequick.com",
                        "auditor": "https://qa-auditor.codequick.com",
                    }
                elif request.headers['Origin'] in ['https://staging-manager.codequick.com', 'https://staging-qa.codequick.com', 'https://staging-client.codequick.com', 'https://staging-auditor.codequick.com']:
                    front_end_url = {
                        "manager": "https://staging-manager.codequick.com",
                        "qa": "https://staging-qa.codequick.com",
                        "client": "https://staging-client.codequick.com",
                        "auditor": "https://staging-auditor.codequick.com",
                    }
                else:
                    front_end_url = {
                        "client": "http://localhost:3000"
                    }

                url = f"{front_end_url['client']}/auth/reset-password?uid={uid}&token={token}&email={user.email}/"
                reset_url = "<a href =" + url + "> Register </a>"

                html_body = """
                    <b><h1 style="color:#023b93;">Welcome %s %s!</h1>
                    </b>
                    <p style="font-size:15px;">Your account has been created successfully %s.</p>
                    <p style="font-size:15px;">We request you to click on the link below to reset your password and get started</p>
                        %s<br>
                        <br>
                    """ %(
                        user.first_name.title(),
                        user.last_name.title(),
                        user.email,
                        reset_url,
                    )

                plain_body = """
                    <b>
                    <h1 style="color:#023b93;">Welcome %s %s!</h1>
                    <br><br>
                    <p>Your account has been created successfully %s.</p>
                        We request you to click on the link below to reset your password and get started<br>
                        %s<br>
                        <br>
                    """ %(
                        user.first_name.title(),
                        user.last_name.title(),
                        user.email,
                        reset_url
                    )

                send_email(
                    subject="Welcome to CodeQuick",
                    htmlBody=html_body,
                    plainBody = plain_body,
                    to=[user.email,],
                    bcc=["vishnu.kumar@buildingblocks.la", "sankavi.boopathy@buildingblocks.la", "vishnu@mailinator.com"]
                )

        if insurance:
            if insurance['exist_ids']:
                [hospital_.insurance.add(ins_id) for ins_id in insurance['exist_ids']]
 
            if insurance['new_ins_name']:
                for new_ins_ in insurance['new_ins_name']:
                    data, new_ins = Insurance.objects.get_or_create(name=new_ins_)
                    hospital_.insurance.add(data)

        if ehr:
            if ehr['exist_ids']:
                [hospital_.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]
 
            if ehr['new_ehr_name']:
                for new_ehr_ in ehr['new_ehr_name']:
                    data, new_ehr = Ehr.objects.get_or_create(name=new_ehr_)
                    hospital_.ehr.add(data)

        serializer = self.get_serializer(hospital_)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    # Partial Update Yet to do:
    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = Hospital.objects.get(id=kwargs['pk'])
        name = request.data.get("name", '').strip()
        address = request.data.get("address", '').strip()
        patients_per_month = request.data.get("patients_per_month", 0)
        specialties = request.data.get("specialties", None)
        insurance = request.data.get("insurance", None)
        ehr = request.data.get("ehr", None)
        health_system = request.data.get("health_system", None)

        # Validations:
        errors = {}

        if request.data.get('account_contact'):    
            for index, data_ in enumerate(request.data['account_contact']):
                errors[index] = {}
                first_name = data_.get("first_name", '').strip()
                last_name = data_.get("last_name", '').strip()


                if first_name == "":
                        errors[index].update({'first_name': 'Firstname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', first_name):
                    errors[index].update({'first_name': 'Firstname can contain only alphabets'})


                if last_name == "":
                    errors[index].update({'last_name': 'Lastname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', last_name):
                    errors[index].update({'last_name': 'Lastname can contain only alphabets'})


                if data_.get("id")==None:
                    email = data_.get('email', '').strip()
                    if CqUser.objects.filter(email__iexact=email).exists():
                        errors[index].update({"Account_contact":"User with this email already exist"})

                if not errors[index]:
                    errors.clear()

        if name:
            if not re.search('^[A-Za-z ]+$', name):
                errors.update({'name': 'Hospital name can contain only alphabets'})


        if health_system:
            if not HealthSystem.objects.filter(id=health_system):
                errors.update({"health_system": "Given Health System is not present in the table."})

        if specialties:
            if specialties['add_spec_ids'] or specialties['remove_spec_ids']:
                for id_ in specialties['add_spec_ids'] + specialties['remove_spec_ids']:
                    if not Specialty.objects.filter(id=id_).exists():
                        errors.update({"specialty": f"Given Specialty {id_} is not present."})            

        if insurance:
            if insurance['exist_ids']:
                for id_ in insurance['exist_ids']:
                    if not Insurance.objects.filter(id=id_).exists():
                        errors.update({"insurance": f"Given Insurance {id_} is not present."})

        if ehr:
            if ehr['exist_ids']:
                for id_ in ehr['exist_ids']:
                    if not Ehr.objects.filter(id=id_).exists():
                        errors.update({"ehr": f"Given Ehr {id_} is not present."})

        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        try:
            if Hospital.objects.filter(name__iexact=name).exclude(name=name).exists():
                errors.update({'name': 'Hospital with this name already exists.'})
            else:
                Hospital.objects.filter(id=instance.id).update(name=name)

        except Exception as e:
            return Response({"name": "Hospital with this name already exists."}, status=status.HTTP_400_BAD_REQUEST)

        health_system_ = HealthSystem.objects.get(id=health_system, is_active=True)
        # hospital_ = Hospital.objects.create(name=name, address=address, patients_per_month=patients_per_month, health_system=health_system_, is_active=True)
        # department_ = Department.objects.get(hospital=hospital_)

        if request.data.get('account_contact'):
            for data_ in request.data['account_contact']:
                if data_.get("id") == None:
                    user = CqUser.objects.create(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'], role="CLIENT")
                    client_user = Client.objects.create(user=user, user_type="HOSPITAL", is_primary=data_['is_primary'])
                    instance.spoc.add(client_user)

                    uid = urlsafe_base64_encode(force_bytes(user.id))
                    token = PasswordResetTokenGenerator().make_token(user)

                    if request.headers['Origin'] in ['https://dev-manager.codequick.com', 'https://dev-client.codequick.com',]:
                        front_end_url = {
                            "client": "https://dev-client.codequick.com",
                        }

                    elif request.headers['Origin'] in ['https://qa-manager.codequick.com', 'https://qa-qa.codequick.com', 'https://qa-client.codequick.com', 'https://qa-auditor.codequick.com']:
                        front_end_url = {
                            "manager": "https://qa-manager.codequick.com",
                            "qa": "https://qa-qa.codequick.com",
                            "client": "https://qa-client.codequick.com",
                            "auditor": "https://qa-auditor.codequick.com",
                        }

                    elif request.headers['Origin'] in ['https://staging-manager.codequick.com', 'https://staging-qa.codequick.com', 'https://staging-client.codequick.com', 'https://staging-auditor.codequick.com']:
                        front_end_url = {
                            "manager": "https://staging-manager.codequick.com",
                            "qa": "https://staging-qa.codequick.com",
                            "client": "https://staging-client.codequick.com",
                            "auditor": "https://staging-auditor.codequick.com",
                        }

                    else:
                        front_end_url = {
                            "client": "http://localhost:3000"
                        }

                    url = f"{front_end_url['client']}/auth/reset-password?uid={uid}&token={token}&email={user.email}/"
                    reset_url = "<a href =" + url + "> Register </a>"

                    html_body = """
                        <b><h1 style="color:#023b93;">Welcome %s %s!</h1>
                        </b>
                        <p style="font-size:15px;">Your account has been created successfully %s.</p>
                        <p style="font-size:15px;">We request you to click on the link below to reset your password and get started</p>
                            %s
                        """ %(
                            user.first_name.title(),
                            user.last_name.title(),
                            user.email,
                            reset_url,
                        )

                    plain_body = """
                        <b>
                        <h1 style="color:#023b93;">Welcome %s %s!</h1>
                        <br><br>
                        <p>Your account has been created successfully %s.</p>
                            We request you to click on the link below to reset your password and get started<br>
                            %s<br>
                            <br>
                        """ %(
                            user.first_name.title(),
                            user.last_name.title(),
                            user.email,
                            reset_url
                        )

                    send_email(
                        subject="Welcome to CodeQuick",
                        htmlBody=html_body,
                        plainBody = plain_body,
                        to=[user.email,],
                        bcc=["vishnu.kumar@buildingblocks.la", "sankavi.boopathy@buildingblocks.la", "vishnu@mailinator.com"]
                    )
                else:
                    CqUser.objects.filter(id=data_['id']).update(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'])
                    Client.objects.filter(user__id=data_['id']).update(is_primary=data_['is_primary'])
                    user = CqUser.objects.get(id=data_['id'])

                if specialties:
                    if specialties['add_spec_ids']:
                        [instance.specialty.add(each_) for each_ in specialties['add_spec_ids']]
                        # for user_ in request.data['account_contact']:
                        #     user = CqUser.objects.get(email=user_['email'])
                            # [user.specialties.add(each_) for each_ in specialties['add_spec_ids']]

                    if specialties['remove_spec_ids']:
                        #[instance.specialty.remove(each_) for each_ in specialties['remove_spec_ids']]
                        Department.objects.filter(hospital=instance, specialty__id__in=specialties['remove_spec_ids']).update(is_deleted=True, is_active=False)
                        # for user_ in request.data['account_contact']:
                        #     user = CqUser.objects.get(email=user_['email'])
                            # [user.specialties.remove(each_) for each_ in specialties['remove_spec_ids']]     
                        [instance.specialty.remove(each_) for each_ in specialties['remove_spec_ids']]

                    if specialties['new_spec_name']:
                        for new_spec_ in specialties['new_spec_name']:
                            data, new_spec = Specialty.objects.get_or_create(name=new_spec_)
                            instance.specialty.add(data)
                            # for user_ in request.data['account_contact']:
                            #     user = CqUser.objects.get(email=user_['email'])
                                # user.specialties.add(data)

        if insurance:
            instance.insurance.clear()
            if insurance['exist_ids']:
                [instance.insurance.add(ins_id) for ins_id in insurance['exist_ids']]
 
            if insurance['new_ins_name']:
                for new_ins_ in insurance['new_ins_name']:
                    data, new_ins = Insurance.objects.get_or_create(name=new_ins_)
                    instance.insurance.add(data)

        if ehr:
            instance.ehr.clear()
            if ehr['exist_ids']:
                [instance.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]
 
            if ehr['new_ehr_name']:
                for new_ehr_ in ehr['new_ehr_name']:
                    data, new_ehr = Ehr.objects.get_or_create(name=new_ehr_)
                    instance.ehr.add(data)

        instance.name = name
        instance.address = address
        instance.patients_per_month = patients_per_month
        instance.save()

        serializer = self.get_serializer(instance)        
        return Response(serializer.data, status=status.HTTP_200_OK)

    def destroy(self, request, pk):
        instance = self.get_object()
        instance.is_active = False
        instance.is_deleted = True

        instance.save()
        return Response("Hospital deleted successfully")

    @action(detail=False, methods=['get',], url_path="validation")
    def validation(self, request):
        queryset = Hospital.objects.filter(is_deleted=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class DepartmentViewSet(SerializerClassMixin, viewsets.ModelViewSet):
    queryset = Department.objects.filter(is_deleted=False)
    serializer_class = DepartmentSerializer
    serializer_action_classes = {
        "providers": ProviderSerializer,
        "department_dropdown": HospitalDepartmentDropDownSerializer,
        "validation": DepartmentValidationSerializer,
        "hospital_department": ProviderSerializer,
    }
    permission_classes = (permissions.IsAuthenticated,)

    def list(self, request):
        ordering = self.request.query_params.get('ordering') if self.request.query_params.get('ordering') else "id"
        queryset = self.get_queryset().order_by(ordering)

        if hospital:=request.query_params.get('hospital'):
            queryset = queryset.filter(hospital=hospital).order_by(ordering)

        if search := self.request.query_params.get('search'):
            queryset = queryset.filter(specialty__name__istartswith=search)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # Passing hospital id in pk:
    @transaction.atomic
    @action(detail=True, methods=['post'])
    def post(self, request, pk):
        name = request.data.get("name", '').strip()
        address = request.data.get("address", '').strip()
        insurance = request.data.get("insurance", None)
        ehr = request.data.get("ehr", None)
        department_id = request.data.get("department_id", None)
        hospital_name = request.data.get("hospital_name", '').strip()

        # Validations:
        errors = {}

        if request.data.get('account_contact'):
            for index, data_ in enumerate(request.data['account_contact']):
                errors[index] = {}
                first_name = data_.get("first_name", '').strip()
                last_name = data_.get("last_name", '').strip()
                email = data_.get("email", '').strip()


                if first_name == "":
                    errors[index].update({'first_name': 'Firstname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', first_name):
                    errors[index].update({'first_name': 'Firstname can contain only alphabets'})


                if last_name == "":
                    errors[index].update({'last_name': 'Lastname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', last_name):
                    errors[index].update({'last_name': 'Lastname can contain only alphabets'})


                if email:

                    if email == "":
                        errors[index].update({'email': 'Email cannot be blank'})

                    if CqUser.objects.filter(email__iexact=email).exists():
                        errors[index].update({'email': f'Email {email} already taken'})


                if not errors[index]:
                    errors.clear()

        if not Hospital.objects.filter(id=pk).exists():
            errors.update({'pk': 'Given hospital is not present in the table.'})

        if name:
            if not re.search('^[A-Za-z ]+$', name):
                errors.update({'name': 'Departmant name can contain only alphabets'})

            if Specialty.objects.filter(name__icontains=name, type='MANAGER').exists():
                errors.update({'name': 'Department with this name already exists.'})

        if insurance:
            if insurance['exist_ids']:
                for id_ in insurance['exist_ids']:
                    if not Insurance.objects.filter(id=id_).exists():
                        errors.update({"insurance": f"Given Insurance {id_} is not present."})

        if ehr:
            if ehr['exist_ids']:
                for id_ in ehr['exist_ids']:
                    if not Ehr.objects.filter(id=id_).exists():
                        errors.update({"ehr": f"Given Ehr {id_} is not present."})

        if hospital_name:
            try:
                if hospital_name != Hospital.objects.get(id=pk, is_deleted=False).name:
                    errors.update({'hospital_name': "Please give the proper hospital name."})
            except Exception as e:
                return Response({'hospital_name': "Please give the proper hospital name."})

        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        if department_id:
            new_specialty = Specialty.objects.get(id=department_id)
            if Department.objects.filter(hospital__id=pk, specialty=new_specialty):
                return Response({'department': "This department already exists in the given hospital"}, status=status.HTTP_400_BAD_REQUEST)
        else:
            new_specialty, data = Specialty.objects.get_or_create(name=name)


        hospital_ = Hospital.objects.get(id=pk)
        # hospital_.address = address if address else None
        hospital_.save()

        department = Department.objects.create(
            specialty=new_specialty,
            hospital=hospital_,
            goal = hospital_.goal,
            top_parameter = hospital_.top_parameter,
            bottom_parameter = hospital_.bottom_parameter,
            is_active=False
            )

        if request.data.get('account_contact'):
            for data_ in request.data['account_contact']:
                user = CqUser.objects.create(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'], role="CLIENT")
                # user.specialties.add(new_specialty)

                client_user = Client.objects.create(user=user, user_type="DEPARTMENT", is_primary=data_['is_primary'])
                hospital_.spoc.add(client_user)
                department.spoc.add(client_user)


                uid = urlsafe_base64_encode(force_bytes(user.id))
                token = PasswordResetTokenGenerator().make_token(user)

                if request.headers['Origin'] in ['https://dev-manager.codequick.com', 'https://dev-client.codequick.com',]:
                        front_end_url = {
                            "client": "https://dev-client.codequick.com",
                        }

                elif request.headers['Origin'] in ['https://qa-manager.codequick.com', 'https://qa-qa.codequick.com', 'https://qa-client.codequick.com', 'https://qa-auditor.codequick.com']:
                    front_end_url = {
                        "manager": "https://qa-manager.codequick.com",
                        "qa": "https://qa-qa.codequick.com",
                        "client": "https://qa-client.codequick.com",
                        "auditor": "https://qa-auditor.codequick.com",
                    }

                elif request.headers['Origin'] in ['https://staging-manager.codequick.com', 'https://staging-qa.codequick.com', 'https://staging-client.codequick.com', 'https://staging-auditor.codequick.com']:
                    front_end_url = {
                        "manager": "https://staging-manager.codequick.com",
                        "qa": "https://staging-qa.codequick.com",
                        "client": "https://staging-client.codequick.com",
                        "auditor": "https://staging-auditor.codequick.com",
                    }

                else:
                    front_end_url = {
                        "client": "http://localhost:3000"
                    }

                url = f"{front_end_url['client']}/auth/reset-password?uid={uid}&token={token}&email={user.email}/"
                reset_url = "<a href =" + url + "> Register </a>"

                html_body = """
                    <b><h1 style="color:#023b93;">Welcome %s %s!</h1>
                    </b>
                    <p style="font-size:15px;">Your account has been created successfully %s.</p>
                    <p style="font-size:15px;">We request you to click on the link below to reset your password and get started</p>
                        %s<br>
                        <br>
                    """ %(
                        user.first_name.title(),
                        user.last_name.title(),
                        user.email,
                        reset_url,
                    )

                plain_body = """
                    <b>
                    <h1 style="color:#023b93;">Welcome %s %s!</h1>
                    <br><br>
                    <p>Your account has been created successfully %s.</p>
                        We request you to click on the link below to reset your password and get started<br>
                        %s<br>
                        <br>
                    """ %(
                        user.first_name.title(),
                        user.last_name.title(),
                        user.email,
                        reset_url
                    )

                send_email(
                    subject="Welcome to CodeQuick",
                    htmlBody=html_body,
                    plainBody = plain_body,
                    to=[user.email,],
                    bcc=["vishnu.kumar@buildingblocks.la", "sankavi.boopathy@buildingblocks.la", "vishnu@mailiantor.com"]
                )

        if insurance:
            if insurance['exist_ids']:
                [hospital_.insurance.add(ins_id) for ins_id in insurance['exist_ids']]
 
            if insurance['new_ins_name']:
                for new_ins_ in insurance['new_ins_name']:
                    data, new_ins = Insurance.objects.get_or_create(name=new_ins_)
                    hospital_.insurance.add(data)

        if ehr:
            if ehr['exist_ids']:
                [hospital_.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]
 
            if ehr['new_ehr_name']:
                for new_ehr_ in ehr['new_ehr_name']:
                    data, new_ehr = Ehr.objects.get_or_create(name=new_ehr_)
                    hospital_.ehr.add(data)

        serializer = self.get_serializer(department)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = Department.objects.get(id=kwargs['pk'])
        name = request.data.get("name", '').strip()
        address = request.data.get("address", '').strip()
        insurance = request.data.get("insurance", None)
        ehr = request.data.get("ehr", None)
        hospital = request.data.get("hospital", None)
        department_id = request.data.get("department_id", None)

        # Validation:
        errors = {}

        if request.data.get('account_contact'):
            for index, data_ in enumerate(request.data['account_contact']):

                errors[index] = {}
                first_name = data_.get("first_name", '').strip()
                last_name = data_.get("last_name", '').strip()


                if first_name == "":
                    errors[index].update({'first_name': 'Firstname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', first_name):
                    errors[index].update({'first_name': 'Firstname can contain only alphabets'})


                if last_name == "":
                    errors[index].update({'last_name': 'Lastname cannot be blank'})

                elif not re.search('^[A-Za-z ]+$', last_name):
                    errors[index].update({'last_name': 'Lastname can contain only alphabets'})


                if data_.get("id")==None:
                    email = data_.get("email", '').strip()

                    if email == "":
                        errors[index].update({'email': 'Email cannot be blank'})

                    if CqUser.objects.filter(email__iexact=email).exists():
                        errors[index].update({"Account_contact":"User with this email already exist"})

                if not errors[index]:
                    errors.clear()

        if insurance:
            if insurance['exist_ids']:
                for id_ in insurance['exist_ids']:
                    if not Insurance.objects.filter(id=id_).exists():
                        errors.update({"insurance": f"Given Insurance {id_} is not present."})

        if ehr:
            if ehr['exist_ids']:
                for id_ in ehr['exist_ids']:
                    if not Ehr.objects.filter(id=id_).exists():
                        errors.update({"ehr": f"Given Ehr {id_} is not present."})

        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        hospital = Hospital.objects.get(id=hospital)

        if department_id:
            new_specialty = Specialty.objects.get(id=department_id)
        else:
            new_specialty, new_spec = Specialty.objects.get_or_create(defaults={'name': name}, name__iexact=name)

        dept, exist_dept = Department.objects.get_or_create(specialty=new_specialty, hospital=hospital)

        if not exist_dept:
            [dept.spoc.add(Client.objects.get(id=spocs['id'])) for spocs in instance.spoc.values('id')]
            instance.spoc.clear()

        if request.data.get('account_contact'):
            for data_ in request.data.get('account_contact'):
                if data_.get("id") == None:
                    user = CqUser.objects.create(first_name=data_['first_name'], last_name=data_['last_name'], email=data_['email'], role="CLIENT")
                    client_user = Client.objects.create(user=user, user_type="DEPARTMENT", is_primary=data_['is_primary'])
                    dept.spoc.add(client_user)

                    uid = urlsafe_base64_encode(force_bytes(user.id))
                    token = PasswordResetTokenGenerator().make_token(user)

                    if request.headers['Origin'] in ['https://dev-manager.codequick.com', 'https://dev-client.codequick.com',]:
                        front_end_url = {
                            "client": "https://dev-client.codequick.com",
                        }

                    elif request.headers['Origin'] in ['https://qa-manager.codequick.com', 'https://qa-qa.codequick.com', 'https://qa-client.codequick.com', 'https://qa-auditor.codequick.com']:
                        front_end_url = {
                            "manager": "https://qa-manager.codequick.com",
                            "qa": "https://qa-qa.codequick.com",
                            "client": "https://qa-client.codequick.com",
                            "auditor": "https://qa-auditor.codequick.com",
                        }

                    elif request.headers['Origin'] in ['https://staging-manager.codequick.com', 'https://staging-qa.codequick.com', 'https://staging-client.codequick.com', 'https://staging-auditor.codequick.com']:
                        front_end_url = {
                            "manager": "https://staging-manager.codequick.com",
                            "qa": "https://staging-qa.codequick.com",
                            "client": "https://staging-client.codequick.com",
                            "auditor": "https://staging-auditor.codequick.com",
                        }

                    else:
                        front_end_url = {
                            "client": "http://localhost:3000"
                        }

                    url = f"{front_end_url['client']}/auth/reset-password?uid={uid}&token={token}&email={user.email}/"
                    reset_url = "<a href =" + url + "> Register </a>"

                    html_body = """
                        <b><h1 style="color:#023b93;">Welcome %s %s!</h1>
                        </b>
                        <p style="font-size:15px;">Your account has been created successfully %s.</p>
                        <p style="font-size:15px;">We request you to click on the link below to reset your password and get started</p>
                            %s<br>
                            <br>
                        """ %(
                            user.first_name.title(),
                            user.last_name.title(),
                            user.email,
                            reset_url,
                        )

                    plain_body = """
                        <b>
                        <h1 style="color:#023b93;">Welcome %s %s!</h1>
                        <br><br>
                        <p>Your account has been created successfully %s.</p>
                            We request you to click on the link below to reset your password and get started<br>
                            %s<br>
                            <br>
                        """ %(
                            user.first_name.title(),
                            user.last_name.title(),
                            user.email,
                            reset_url
                        )

                    send_email(
                        subject="Welcome to CodeQuick",
                        htmlBody=html_body,
                        plainBody = plain_body,
                        to=[user.email,],
                        bcc=["vishnu.kumar@buildingblocks.la", "sankavi.boopathy@buildingblocks.la", "vishnu@mailinator.com"]
                    )

                else:
                    CqUser.objects.filter(id=data_['id']).update(first_name=data_['first_name'], last_name=data_['last_name'])
                    Client.objects.filter(user__id=data_['id']).update(is_primary=data_['is_primary'])

        if address:
            Hospital.objects.filter(id=hospital.id).update(address=address)

        # hospital_ = Hospital.objects.get(id=hospital.id)

        if insurance:
            hospital.insurance.clear()
            if insurance['exist_ids']:
                [hospital.insurance.add(ins_id) for ins_id in insurance['exist_ids']]
 
            if insurance['new_ins_name']:
                for new_ins_ in insurance['new_ins_name']:
                    data, new_ins = Insurance.objects.get_or_create(name=new_ins_)
                    hospital.insurance.add(data)

        if ehr:
            hospital.ehr.clear()
            if ehr['exist_ids']:
 
                [hospital.ehr.add(ehr_id) for ehr_id in ehr['exist_ids']]
            if ehr['new_ehr_name']:
                for new_ehr_ in ehr['new_ehr_name']:
                    data, new_ehr = Ehr.objects.get_or_create(name=new_ehr_)
                    hospital.ehr.add(data)

        # serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer = self.get_serializer(dept, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(serializer.data, status=status.HTTP_200_OK)

    def destroy(self, request, pk):
        instance = self.get_object()
        instance.is_active = False
        instance.is_deleted = True
        instance.save()
        return Response("Department deleted successfully")

    # To fetch department exclude given hospital_id:
    @action(detail=True, methods=['get',], url_path='department_dropdown')
    def department_dropdown(self, request, pk):

        existing_specialty_ids = [ each_.specialty.id for each_ in self.get_queryset().filter(hospital=pk) ]
        specialty_ids = list(set(existing_specialty_ids))
        queryset = Specialty.objects.filter(type='MANAGER').exclude(id__in=specialty_ids)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['get',], url_path='providers')
    def providers(self, request, pk):
        queryset = Department.objects.get(id=pk).providers.filter(user__is_deleted=False)
        if search := request.query_params.get('search'):
            queryset = queryset.annotate(fullname=Concat('user__first_name', Value(' '), 'user__last_name')).filter(fullname__icontains=search)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get',], url_path="validation")
    def validation(self, request):
        queryset = Department.objects.filter(is_deleted=False)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # To get Hospital Assigned Department, Pass Hospital id in PK:
    @action(detail=True, methods=['get'], url_path='hospital_department')
    def hospital_department(self, request, pk):
        queryset = Department.objects.filter(hospital=pk)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ProviderViewSet(viewsets.ModelViewSet):
    serializer_class = ProviderSerializer
    permission_classes = (permissions.IsAuthenticated,)
    queryset = Client.objects.filter(user_type='PROVIDER', is_deleted=False)

    def list(self, request):
        queryset = self.get_queryset().filter(user__is_active=True)
        ordering = self.request.query_params.get('ordering') if self.request.query_params.get('ordering') else "id"
        search = self.request.query_params.get('search')

        if search:
            queryset = queryset.filter(user__email__istartswith=search).order_by("user__email")

        if ordering == "name":
            queryset = queryset.order_by("user__first_name", "user__last_name")
        elif ordering == "-name":
            queryset = queryset.order_by("-user__first_name", "-user__last_name")
        else:
            queryset = queryset.order_by(ordering)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


    @transaction.atomic
    @action(detail=True, methods=['post'], permission_classes=[permissions.AllowAny, ])
    def post(self, request, pk):
        first_name = request.data.get("first_name", '').strip()
        last_name = request.data.get("last_name", '').strip()
        email = request.data.get("email", '').strip().lower()
        provider_id = request.data.get("provider_id", None)

        department = Department.objects.get(id=pk)
        # Validations:
        errors = {}

        if provider_id:

            if not Client.objects.filter(id=provider_id, user_type="PROVIDER").exists():
                errors.update({'user': "Please give the provider's email."})

            if not Department.objects.filter(id=pk).exists():
                errors.update({'user': "Given Department is not present in the table."})

            if errors:
                return Response(errors, status=status.HTTP_400_BAD_REQUEST)

            provider = Client.objects.get(id=provider_id)
            if provider.department_providers.values():
                return Response({"provider": f"This provider already belongs to {provider.department_providers.values('specialty__name').last()['specialty__name']} department."}, status=status.HTTP_400_BAD_REQUEST)

            department.providers.add(provider)

        else:
            if first_name == "":
                errors.update({'first_name': 'Firstname cannot be blank'})

            elif not re.search('^[A-Za-z ]+$', first_name):
                errors.update({'first_name': 'Firstname can contain only alphabets'})

            if last_name == "":
                errors.update({'last_name': 'Lastname cannot be blank'})

            elif not re.search('^[A-Za-z ]+$', last_name):
                errors.update({'last_name': 'Lastname can contain only alphabets'})

            if email == "":
                errors.update({'email': 'Email cannot be blank'})

            if CqUser.objects.filter(email__iexact=email, is_deleted=False).exists():
                errors.update({'email': 'Email already taken'})


            if errors:
                return Response(errors, status=status.HTTP_400_BAD_REQUEST)

            if CqUser.objects.filter(email__iexact=email, is_deleted=True).exists():
                CqUser.objects.filter(email__iexact=email).update(first_name=first_name, last_name=last_name, role="CLIENT", password='', is_active=True, is_deleted=False)
                user = CqUser.objects.get(email=email)
                provider = Client.objects.get(user=user, user_type="PROVIDER")
                provider.is_deleted=False
                provider.save()

                if pk != str(provider.department_providers.last().id):
                    provider.department_providers.clear()
                    provider.department_providers.add(department)

            else:
                user = CqUser.objects.create(first_name=first_name, last_name=last_name, email=email, role="CLIENT")
                provider = Client.objects.create(user=user, user_type="PROVIDER")

            # user.specialties.add(department.specialty)
            #department.providers.add(provider)
            ProviderStatistics.objects.create(
                provider = provider,
                department = department,
                goal = department.goal,
                top_parameter = department.top_parameter,
                bottom_parameter = department.bottom_parameter
                )

            uid = urlsafe_base64_encode(force_bytes(user.id))
            token = PasswordResetTokenGenerator().make_token(user)

            # if request.headers['Origin'] in ['https://dev-manager.codequick.com', 'https://dev-client.codequick.com',]:
            if 1 < 2:
                front_end_url = {
                    "client": "https://dev-client.codequick.com",
                }
            
            elif request.headers['Origin'] in ['https://qa-manager.codequick.com', 'https://qa-qa.codequick.com', 'https://qa-client.codequick.com', 'https://qa-auditor.codequick.com']:
                front_end_url = {
                    "manager": "https://qa-manager.codequick.com",
                    "qa": "https://qa-qa.codequick.com",
                    "client": "https://qa-client.codequick.com",
                    "auditor": "https://qa-auditor.codequick.com",
                }

            elif request.headers['Origin'] in ['https://staging-manager.codequick.com', 'https://staging-qa.codequick.com', 'https://staging-client.codequick.com', 'https://staging-auditor.codequick.com']:
                front_end_url = {
                    "manager": "https://staging-manager.codequick.com",
                    "qa": "https://staging-qa.codequick.com",
                    "client": "https://staging-client.codequick.com",
                    "auditor": "https://staging-auditor.codequick.com",
                }

            else:
                front_end_url = {
                    "client": "http://localhost:3000"
                }

            url = f"{front_end_url['client']}/auth/reset-password?uid={uid}&token={token}&email={user.email}/"
            reset_url = "<a href =" + url + "> Register </a>"


            html_body = """
                <b><h1 style="color:#023b93;">Welcome %s %s!</h1>
                </b>
                <p style="font-size:15px;">Your account has been created successfully <b>%s</b>.</p>
                <p style="font-size:15px;">We request you to click on the link below to reset your password and get started</p>
                    %s<br>
                    <br>
                """ %(
                    user.first_name.title(),
                    user.last_name.title(),
                    user.email,
                    reset_url,
                )

            plain_body = """
                <b>
                <h1 style="color:#023b93;">Welcome %s %s!</h1>
                <br><br>
                <p>Your account has been created successfully <b>%s</b>.</p>
                    We request you to click on the link below to reset your password and get started<br>
                    %s<br>
                    <br>
                """ %(
                    user.first_name.title(),
                    user.last_name.title(),
                    user.email,
                    reset_url
                )

            send_email(
                subject="Welcome to CodeQuick",
                htmlBody=html_body,
                plainBody = plain_body,
                to=[user.email,],
                bcc=["vishnu.kumar@buildingblocks.la", "vishnu@mailinator.com", "sankavi.boopathy@buildingblocks.la"]
            )
        serializer = self.get_serializer(provider)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = Client.objects.get(id=kwargs['pk'])
        first_name = request.data.get("first_name", '').strip()
        last_name = request.data.get("last_name", '').strip()
        # is_active = request.data.get("is_active", None)

        # Validations:
        errors = {}
        if first_name == "":
            errors.update({'first_name': 'Firstname cannot be blank'})

        elif not re.search('^[A-Za-z ]+$', first_name):
            errors.update({'first_name': 'Firstname can contain only alphabets'})

        if last_name == "":
            errors.update({'last_name': 'Lastname cannot be blank'})

        elif not re.search('^[A-Za-z ]+$', last_name):
            errors.update({'last_name': 'Lastname can contain only alphabets'})

        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        CqUser.objects.filter(client__id=kwargs['pk']).update(first_name=first_name, last_name=last_name)
        # instance.user.first_name = first_name
        # instance.user.last_name = last_name
        # instance.user.is_active = is_active
        # instance.save()

        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()

        return Response(serializer.data, status=status.HTTP_200_OK)


    def destroy(self, request, pk):
        instance = CqUser.objects.get(client__id=pk)
        instance.is_active = False
        instance.is_deleted = True
        instance.client.is_deleted = True
        instance.save()
        return Response({"message": "Provider deleted successfully"}, status=status.HTTP_200_OK)


class InsuranceViewSet(viewsets.ModelViewSet):
    queryset = Insurance.objects.all()
    serializer_class = InsuranceSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def list(self, request):
        queryset = self.get_queryset().order_by("id")
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class EhrViewSet(viewsets.ModelViewSet):
    queryset = Ehr.objects.all()
    serializer_class = EhrSerializer
    permission_classes = (permissions.IsAuthenticated,)

    def list(self, request):
        queryset = self.get_queryset().order_by("id")
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
